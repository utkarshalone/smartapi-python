import requests
import pandas as pd
import time
from datetime import datetime
import os
from openpyxl import load_workbook
import copy





# package import statement
from SmartApi import SmartConnect #or from SmartApi.smartConnect import SmartConnect
import pyotp
from logzero import logger
def generate_session():
    api_key = "WoaxRnYv"
    username = 'U21226'
    pwd = '4444'
    smartApi = SmartConnect(api_key)
    try:
        token = "4WD4HHYWSCEL5GPK7WXKUCPNVM"
        totp = pyotp.TOTP(token).now()
    except Exception as e:
        logger.error("Invalid Token: The provided token is not valid.")
        raise e

    correlation_id = "abcde"
    data = smartApi.generateSession(username, pwd, totp)

    if data['status'] == False:
        logger.error(data)
        
    else:
        # login api call
        # logger.info(f"You Credentials: {data}")
        authToken = data['data']['jwtToken']
        refreshToken = data['data']['refreshToken']
        # fetch the feedtoken
        feedToken = smartApi.getfeedToken()
        # fetch User Profile
        res = smartApi.getProfile(refreshToken)
        smartApi.generateToken(refreshToken)
    return api_key, authToken



def get_scrip_master():
    url = "https://margincalculator.angelbroking.com/OpenAPI_File/files/OpenAPIScripMaster.json"
    df = pd.DataFrame(requests.get(url).json())
    df.columns = df.columns.str.lower()
    return df

def get_atm_tokens(df, spot, window=4):
    df = df[(df["name"] == "NIFTY") & (df["instrumenttype"] == "OPTIDX")]
    df["expiry"] = pd.to_datetime(df["expiry"], format="%d%b%Y", errors="coerce")
    nearest_expiry = df["expiry"].min()
    latest_df = df[df["expiry"] == nearest_expiry].copy()
    latest_df["strike"] = pd.to_numeric(latest_df["strike"], errors="coerce")
    latest_df = latest_df.dropna(subset=["strike"])
    latest_df["strike_diff"] = (latest_df["strike"] - spot).abs()
    atm_strike = latest_df.loc[latest_df["strike_diff"].idxmin(), "strike"]
    strikes_sorted = sorted(latest_df["strike"].unique())
    atm_index = strikes_sorted.index(atm_strike)
    selected_strikes = strikes_sorted[max(0, atm_index-window):atm_index+window+1]
    filtered_df = latest_df[latest_df["strike"].isin(selected_strikes)] 
    tokens = sorted(filtered_df["token"].tolist())
    return tokens, filtered_df["strike"].tolist(), atm_strike, selected_strikes

def get_nifty_option_tokens(df, spot, window=4):
    # Filter only NIFTY OPTIDX
    df = df[(df["name"] == "NIFTY") & (df["instrumenttype"] == "OPTIDX")]
    df["expiry"] = pd.to_datetime(df["expiry"], format="%d%b%Y", errors="coerce")
    nearest_expiry = df["expiry"].min()

    # Work only on nearest expiry
    latest_df = df[df["expiry"] == nearest_expiry].copy()
    latest_df["strike"] = pd.to_numeric(latest_df["strike"], errors="coerce")
    latest_df = latest_df.dropna(subset=["strike"])
    latest_df["strike_diff"] = (latest_df["strike"] - spot).abs()

    # ATM strike
    atm_strike = latest_df.loc[latest_df["strike_diff"].idxmin(), "strike"]
    strikes_sorted = sorted(latest_df["strike"].unique())
    atm_index = strikes_sorted.index(atm_strike)
    selected_strikes = strikes_sorted[max(0, atm_index - window): atm_index + window + 1]

    # ATM ± window filter
    filtered_df = latest_df[latest_df["strike"].isin(selected_strikes)]

    # ✅ Return all tokens and symbols (from latest_df, not filtered)
    all_tokens = sorted(latest_df["token"].tolist())
    all_symbols = sorted(latest_df["symbol"].tolist())

    # ✅ Also return ATM ± window tokens and symbols
    filtered_tokens = sorted(filtered_df["token"].tolist())
    filtered_symbols = sorted(filtered_df["symbol"].tolist())

    return {
        "all_tokens": all_tokens,
        "all_symbols": all_symbols,
        "atm_strike": atm_strike,
        "selected_strikes": selected_strikes,
        "filtered_tokens": filtered_tokens,
        "filtered_symbols": filtered_symbols
    }



def fetch_batch(token_list, api_key, authToken):
    print("Fetching batch from fetch_batch:", token_list)
    payload = {
        "mode": "FULL",
        "exchangeTokens": {"NFO": token_list}
    }
    headers = {
        "X-PrivateKey": api_key,
        "Accept": "application/json",
        "X-SourceID": "WEB",
        "X-UserType": "USER",
        "Authorization": f"{authToken}",
        "Content-Type": "application/json"
    }
    resp = requests.post(
        "https://apiconnect.angelone.in/rest/secure/angelbroking/market/v1/quote/",
        json=payload,
        headers=headers
    )
    return resp.json()



def dicts_of_dfs_equal(dict1, dict2):
    if dict1.keys() != dict2.keys():
        return False
    for key in dict1:
        if len(dict1[key]) != len(dict2[key]):
            return False
        for df1, df2 in zip(dict1[key], dict2[key]):
            if not df1.equals(df2):
                return False
    return True


def write_to_excel(excel_filename, symbol_data_dict):
    columns_to_keep = [
        "tradingSymbol", "ltp", "open", "high", "low", "close",
        "netChange", "percentChange", "avgPrice", "tradeVolume",
        "opnInterest", "exchFeedTime", "fetch_time"
    ]

    if os.path.exists(excel_filename):
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as excel_writer:
            book = load_workbook(excel_filename)
            for symbol, df_list in symbol_data_dict.items():
                df_concat = pd.concat(df_list, ignore_index=True)
                available_columns = [col for col in columns_to_keep if col in df_concat.columns]
                df_concat = df_concat[available_columns]
                sheet_name = str(symbol)[:31]
                if sheet_name in book.sheetnames:
                    existing_df = pd.read_excel(excel_filename, sheet_name=sheet_name)
                    df_concat = pd.concat([existing_df, df_concat], ignore_index=True)
                df_concat.to_excel(excel_writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(excel_filename, engine="xlsxwriter") as excel_writer:
            for symbol, df_list in symbol_data_dict.items():
                df_concat = pd.concat(df_list, ignore_index=True)
                available_columns = [col for col in columns_to_keep if col in df_concat.columns]
                df_concat = df_concat[available_columns]
                sheet_name = str(symbol)[:31]
                df_concat.to_excel(excel_writer, sheet_name=sheet_name, index=False)

def main():
    spot = 2450000  # Replace with actual spot value if needed
    BATCH_SIZE = 25  # Define batch size for fetching dat
    api_key, authToken = generate_session()  # Ensure session is generated before proceeding
    df = get_scrip_master()
    # tokens, strikes, atm_strike, selected_strikes = get_atm_tokens(df, spot)   #, symbol

    # print(f"ATM strike: {atm_strike}")
    # print(f"Selected tokens (ATM ±4): {tokens}")
    # print("Selected strikes:", strikes)
    # # print(f"Selected symbols:", {symbol})
    # print(f"Selected strikes (ATM ±4): {sorted(selected_strikes)}")

    result = get_nifty_option_tokens(df, spot)

    # print("✅ All tokens:", result["all_tokens"])
    # print("✅ All symbols:", result["all_symbols"])
    # print("✅ ATM Strike:", result["atm_strike"])
    print("✅ Selected strikes:", result["selected_strikes"])
    print("✅ Filtered (ATM ±4) tokens:", result["filtered_tokens"])
    print("✅ Filtered (ATM ±4) symbols:", result["filtered_symbols"])
    tokens = result["filtered_tokens"]
    # tokens = sorted(result["filtered_tokens"].tolist())
    strikes = result["selected_strikes"]
    atm_strike = result["atm_strike"]

    excel_filename = f"oi_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    symbol_data_dict = {}
    # print(latest_df.head())

    # last_fetch_minute = None
    # while True:
    #     now = datetime.now()
    #     # Wait until 8:55 AM to start
    #     if now.hour < 8 or (now.hour == 8 and now.minute < 55):
    #         print("Waiting for 8:55 AM to start data collection...")
    #         # Sleep until 8:55 AM
    #         target = now.replace(hour=8, minute=55, second=0, microsecond=0)
    #         if now >= target:
    #             # Already past 8:55, continue loop
    #             time.sleep(5)
    #         else:
    #             seconds_to_wait = (target - now).total_seconds()
    #             print(f"Sleeping for {int(seconds_to_wait)} seconds until 8:55 AM...")
    #             time.sleep(seconds_to_wait)
    #         continue
    #     # Only fetch if minute is a multiple of 3, second is 2, and not already fetched for this minute
    #     if now.second == 2 and last_fetch_minute != now.minute: #and now.minute % 3 == 0:
    #         last_fetch_minute = now.minute
    #         fetch_time = now.strftime('%Y-%m-%d %H:%M:%S')
    #         print(f"Fetching data at {fetch_time}")
    #         for i in range(0, len(tokens), BATCH_SIZE):
    #             batch = tokens[i:i+BATCH_SIZE]
    #             print(f"  Batch {i//BATCH_SIZE + 1} ({len(batch)} tokens)")
    #             data = fetch_batch(batch, api_key, authToken)
    #             if "data" in data and "fetched" in data["data"]:
    #                 fetched = data["data"]["fetched"]
    #                 items = fetched.items() if isinstance(fetched, dict) else [
    #                     (
    #                         entry.get("tradingSymbol") or entry.get("tradingsymbol") or entry.get("symboltoken") or entry.get("symbol") or f"symbol_{i}",
    #                         entry
    #                     ) for entry in fetched
    #                 ] if isinstance(fetched, list) else []
    #                 for symbol, symbol_data in items:
    #                     df_symbol = pd.json_normalize(symbol_data)
    #                     df_symbol['fetch_time'] = fetch_time
    #                     last_df = symbol_data_dict[symbol][-1] if symbol in symbol_data_dict and symbol_data_dict[symbol] else None
    #                     is_new = True
    #                     if last_df is not None:
    #                         for col in ["opnInterest", "ltp"]:
    #                             if col in df_symbol.columns and col in last_df.columns:
    #                                 if df_symbol[col].iloc[0] == last_df[col].iloc[0]:
    #                                     is_new = False
    #                                 else:
    #                                     is_new = True
    #                                     break
    #                     if is_new:
    #                         symbol_data_dict.setdefault(symbol, []).append(df_symbol)
    #         write_to_excel(excel_filename, symbol_data_dict)
    #         print(f"Data written to Excel file: {excel_filename}")
    #         # Wait until second != 2 to avoid duplicate fetches
    #         while datetime.now().second == 2:
    #             time.sleep(0.2)
    #     else:
    #         time.sleep(0.2)

    new_symbol_dict = {}

    while True:
        for i in range(0, len(tokens), BATCH_SIZE):
            batch = tokens[i:i + BATCH_SIZE]
            print(f"  Batch {i // BATCH_SIZE + 1} ({len(batch)} tokens)")

            data = fetch_batch(batch, api_key, authToken)

            if "data" in data and "fetched" in data["data"]:
                fetched = data["data"]["fetched"]
                items = (
                    fetched.items() if isinstance(fetched, dict) else [
                        (
                            entry.get("tradingSymbol") or entry.get("tradingsymbol") or entry.get("symboltoken") or entry.get("symbol") or f"symbol_{i}",
                            entry
                        ) for entry in fetched
                    ]
                ) if isinstance(fetched, list) or isinstance(fetched, dict) else []

                for symbol, symbol_data in items:
                    df_symbol = pd.json_normalize(symbol_data)
                    fetch_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    df_symbol['fetch_time'] = fetch_time

                    last_df = symbol_data_dict[symbol][-1] if symbol in symbol_data_dict and symbol_data_dict[symbol] else None
                    is_new = True
                    if last_df is not None:
                        for col in ["opnInterest", "ltp"]:
                            if col in df_symbol.columns and col in last_df.columns:
                                if df_symbol[col].iloc[0] == last_df[col].iloc[0]:
                                    is_new = False
                                else:
                                    is_new = True
                                    break

                    if is_new:
                        symbol_data_dict.setdefault(symbol, []).append(df_symbol)

            # Perform deep comparison
            if not dicts_of_dfs_equal(symbol_data_dict, new_symbol_dict):
                print("New data found, writing to Excel...")
                write_to_excel(excel_filename, symbol_data_dict)
                print(f"Data written to Excel file: {excel_filename}")
                new_symbol_dict = copy.deepcopy(symbol_data_dict)
            else:
                print("No new data found, skipping write to Excel.")
        # Wait for 3 seconds before the next fetch
        time.sleep(3)
if __name__ == "__main__":
    main()
